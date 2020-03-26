import hashlib

try:
    import pickle
except ImportError:
    import cPickle as pickle

from django.conf import settings
from django.contrib.sites.models import Site
from django.urls import reverse
from django.http import HttpResponse
from django.utils.encoding import escape_uri_path

from tq_website import settings as my_settings

from post_office import mail, models as post_office_models
import logging

from survey.models import Answer
from courses.models import Course

from django.utils import translation

log = logging.getLogger('tq')

SALT = b"lkd$lrn&"


def calc_checksum(id_str):
    checksum = hashlib.md5()
    checksum.update(id_str.encode())
    checksum.update(SALT)
    return checksum.hexdigest()[:4].lower()


def encode_id(id):
    id_str = str(id)
    return id_str, calc_checksum(id_str)


def decode_id(id_str, checksum):
    try:
        id = int(id_str)
    except ValueError:
        log.warning(
            "The id of survey link with id='{}' and checksum='{}' could not be decoded".format(id_str, checksum))
        return False
    real_checksum = calc_checksum(id_str)  # use the string here, not the int
    if real_checksum != checksum.lower():
        log.warning(
            "The checksum of survey link with id='{}' and checksum='{}' does not match the real checksum='{}'".format(
                id_str, checksum, real_checksum))
        return False
    return int(id_str)


def create_url(survey_inst):
    id_str, c = encode_id(survey_inst.id)
    return"{}?id={}&c={}".format(reverse('survey:survey_invitation'), escape_uri_path(id_str),
                                   escape_uri_path(c))


def create_full_url(survey_inst):
    return"https://{}{}".format(Site.objects.get(id=settings.SITE_ID).domain, create_url(survey_inst))


def send_invitation(survey_inst):
    if survey_inst.invitation_sent:
        return False

    lang = translation.get_language()
    translation.activate('de')
    url_de = create_full_url(survey_inst)
    translation.activate('en')
    url_en = create_full_url(survey_inst)
    translation.activate(lang)

    context = {
        'first_name': survey_inst.user.first_name,
        'last_name': survey_inst.user.last_name,
        'course': survey_inst.course.type.name,
        'offering': survey_inst.course.offering.name,
        'expires': survey_inst.url_expire_date,
        'url': url_de,
        'url_en': url_en,
        'url_de': url_de,
    }

    template = 'survey_invitation'
    if survey_inst.email_template:
        template = survey_inst.email_template

    if _email_helper(survey_inst.user.email, template, context):
        survey_inst.invitation_sent = True
        survey_inst.save()
        return True
    else:
        return False


def _email_helper(email, template, context):
    """Sending facility. Catches errors due to not existent template."""
    try:
        mail.send(
            [email],
            my_settings.DEFAULT_FROM_EMAIL,
            template=template,
            context=context,
        )
        return True
    except post_office_models.EmailTemplate.DoesNotExist as e:
        log.error("Email Template missing with name: {}".format(template))
        return False


import zipfile
from io import BytesIO

import openpyxl


# exports the subscriptions of course with course_id to fileobj (e.g. a HttpResponse)
def export_surveys(surveys):
    def create_xlsx_sheet(wb, survey, instances, title):
        ws = wb.create_sheet(title=title[:28])

        row_num = 0

        # first fill questions in list to ensure consistent order
        questions = []
        for g in survey.questiongroup_set.all():
            questions += list(g.question_set.all())

        columns = []
        for q in questions:
            columns.append(q.name)

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num]
            font = c.font.copy()
            font.bold = True
            c.font = font

        for inst in instances:
            # only take the newest answer for all questions
            answers = Answer.objects.filter(survey_instance__survey=survey,
                                                   survey_instance__user=inst.user).order_by('-id')

            row_num += 1
            for col_num in range(len(columns)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                res = answers.filter(question=questions[col_num])
                c.value = res[0].text if res else ""

                alignment = c.alignment.copy()
                alignment.wrap_text = True
                c.alignment = alignment

    zipped_file = BytesIO()  # since Python3, this must by BytesIO (not StringIO) since zipfile operates still on Bytes
    with zipfile.ZipFile(zipped_file, 'w') as f:
        for survey in surveys:
            fileobj = BytesIO()    # since Python3, this must by BytesIO (not StringIO) since zipfile operates still on Bytes

            #####################################
            survey_inst = survey.survey_instances.values('course').distinct()
            wb = openpyxl.Workbook()
            if survey_inst.count():
                # remove preinitialized sheet (only if we later add same, no sheets in the end lead to error!)
                wb.remove_sheet(wb.get_active_sheet())

            for d in survey_inst:
                course_id = d['course']  # get the courses out of dict returned by values
                if course_id:
                    course = Course.objects.get(pk=course_id)
                    create_xlsx_sheet(wb, survey,
                                      survey.survey_instances.exclude(last_update=None).filter(course=course_id).all(),
                                      course.name)
                else:  # course_id is None
                    create_xlsx_sheet(wb, survey, survey.survey_instances.exclude(last_update=None).all(),
                                      "COURSE UNSPECIFIC")

            wb.save(fileobj)
            #####################################

            f.writestr(u'Surveys/{}.xlsx'.format(survey.name),
                       fileobj.getvalue())
            fileobj.seek(0)

    content_length = zipped_file.tell()
    zipped_file.seek(0)
    response = HttpResponse(zipped_file, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=Surveys.zip'
    response['Content-Length'] = content_length

    return response
