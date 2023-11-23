from collections import defaultdict

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from . import clean_filename


def export_excel(title, data, has_column_headings=True, multiple=False):
    workbook = Workbook()

    if multiple:
        sheets = dict()
        workbook.remove_sheet(workbook.active)
        for item in data:
            sheet_title = clean_filename(item["name"])[:30]
            sheets[workbook.create_sheet(title=sheet_title)] = item["data"]
    else:
        sheets = {workbook.active: data}

    for sheet, values in sheets.items():
        column_widths = defaultdict(int)
        for row_number, row in enumerate(values, 1):  # ,1 to start at 1
            for col_number, value in enumerate(row, 1):
                try:
                    cell = sheet.cell(row=row_number, column=col_number, value=value)
                except ValueError:
                    cell = sheet.cell(
                        row=row_number, column=col_number, value=str(value)
                    )

                cell.alignment = Alignment(wrapText=True, vertical="center")

                column_widths[col_number] = max(
                    max([len(v) for v in str(value).split("\n")]),
                    column_widths[col_number],
                )

                if has_column_headings and row_number == 1:
                    font = cell.font.copy()
                    font.bold = True
                    cell.font = font

        for col_number, column_width in column_widths.items():
            sheet.column_dimensions[get_column_letter(col_number)].width = (
                column_width * 1.25
            )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename={}.xlsx".format(
        clean_filename(title)
    )

    workbook.save(response)
    return response
