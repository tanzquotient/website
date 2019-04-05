from django.http import HttpResponse
from openpyxl import Workbook

from courses.utils import clean_filename


def export_excel(title, data, has_column_headings=True, multiple=False):

    workbook = Workbook()

    if multiple:
        sheets = dict()
        workbook.remove_sheet(workbook.get_active_sheet())
        for item in data:
            sheet_title = clean_filename(item['name'])[:30]
            sheets[workbook.create_sheet(title=sheet_title)] = item['data']
    else:
        sheets = {workbook.get_active_sheet(): data}

    for sheet, values in sheets.items():
        for row_number, row in enumerate(values):
            for col_number, value in enumerate(row):
                cell = sheet.cell(row=row_number + 1, column=col_number + 1, value=value)
                if has_column_headings and row_number == 0:
                    font = cell.font.copy()
                    font.bold = True
                    cell.font = font

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(clean_filename(title))

    workbook.save(response)
    return response
